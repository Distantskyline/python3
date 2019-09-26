import openpyxl
abc = openpyxl.Workbook()
abc.save('abc.xlsx')


import time
import openpyxl

# 实例化:打开文件
workbook = openpyxl.load_workbook('monitor.xlsx')
# 获取活动表格,默认选择你退出时所在的表格
sheet = workbook.active
# A1代表要修改/写入的坐标,
sheet['A1'] = time.strftime('%Y%m%d%H%M%S', time.localtime())
# 在写入数据的最后要去完成所做修改的保存
workbook.save('monitor.xlsx')




# One: 新建文件
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet['A1'] = time.strftime('%Y%m%d%H%M%S', time.localtime())
workbook.save('monitor1.xlsx')

# Two: 打开已经存在的文件
workbook = openpyxl.load_workbook('monitor2.xlsx')
sheet = workbook.active
sheet['A1'] = time.strftime('%Y%m%d%H%M%S', time.localtime())
workbook.save('monitor2.xlsx')




# 方式一: 直接创建新的sheet来添加数据
workbook = openpyxl.load_workbook('monitor.xlsx')
sheet = workbook.create_sheet(title='nginxLog', index=0)
sheet['A1'] = 'nginxLog'

workbook.save('monitor.xlsx')

# 方式二: 指定打开
workbook = openpyxl.load_workbook('monitor.xlsx')
sheet = workbook['nginxLog']
sheet['B1'] = 'nginxLog'

workbook.save('monitor.xlsx')




workbook = openpyxl.load_workbook('monitor.xlsx')
sheet = workbook.create_sheet(title='nginxLog', index=0)
sheet['A1'] = 'nginxLog'
sheet.append(['date, cpu, memory, disk'])
sheet.append([2018813142512, 65, 40, 60])

workbook.save('monitor.xlsx')